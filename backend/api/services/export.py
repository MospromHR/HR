from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from api.schemas.analytics import (
    AdminStatsResponse,
    CompanyStatsResponse,
    EducationStatsResponse,
)


def _auto_fit_columns(sheet) -> None:
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            value = cell.value
            if value is None:
                continue
            cell_length = len(str(value))
            if cell_length > max_length:
                max_length = cell_length
        sheet.column_dimensions[column_letter].width = max_length + 2


def _write_table(
    sheet,
    headers: Sequence[str],
    rows: Iterable[Sequence[object]],
    column_formats: Sequence[str | None] | None = None,
) -> None:
    sheet.append(list(headers))
    header_font = Font(bold=True)
    for cell in sheet[1]:
        cell.font = header_font

    for row in rows:
        sheet.append(list(row))
        if column_formats:
            for idx, fmt in enumerate(column_formats, start=1):
                if fmt and sheet.cell(row=sheet.max_row, column=idx).value is not None:
                    sheet.cell(row=sheet.max_row, column=idx).number_format = fmt

    _auto_fit_columns(sheet)


def _build_workbook() -> Workbook:
    wb = Workbook()
    # Remove the default sheet and create our own to control titles
    default_sheet = wb.active
    wb.remove(default_sheet)
    return wb


def _summary_sheet(wb: Workbook, title: str, rows: Sequence[tuple[str, object, str | None]]) -> None:
    sheet = wb.create_sheet(title=title)
    sheet.append(["Показатель", "Значение"])
    header_font = Font(bold=True)
    sheet["A1"].font = header_font
    sheet["B1"].font = header_font

    for metric, value, number_format in rows:
        display = value if value is not None else "-"
        sheet.append([metric, display])
        if value is not None and number_format:
            sheet.cell(row=sheet.max_row, column=2).number_format = number_format

    _auto_fit_columns(sheet)


def _enum_to_value(value: object) -> object:
    if hasattr(value, "value"):
        return getattr(value, "value")
    return value


def build_company_stats_report(stats: CompanyStatsResponse) -> bytes:
    wb = _build_workbook()
    _summary_sheet(
        wb,
        "Сводка",
        [
            ("Среднее время закрытия вакансий (дни)", stats.average_vacancy_closure_days, "0.00"),
            ("Принятые студенты", stats.accepted_students, None),
            ("Успешность найма", stats.success_rate, "0.00%"),
            ("Всего заявок", stats.total_applications, None),
            (
                "Среднее число заявок на вакансию",
                stats.average_applications_per_vacancy,
                "0.00",
            ),
            ("Среднее время отклика (дни)", stats.average_response_time_days, "0.00"),
            ("Опубликованные вакансии", stats.published_vacancies, None),
            ("Открытые вакансии", stats.open_vacancies, None),
            ("Текущие утвержденные стажировки", stats.current_approved_internships, None),
            (
                "Средняя длительность утвержденных стажировок (дни)",
                stats.approved_internship_average_duration_days,
                "0.00",
            ),
        ],
    )

    status_sheet = wb.create_sheet(title="Заявки по статусу")
    _write_table(
        status_sheet,
        ["Статус", "Количество"],
        ((_enum_to_value(item.status), item.count) for item in stats.application_status_breakdown),
    )

    engagement_sheet = wb.create_sheet(title="Сотрудничества по статусу")
    _write_table(
        engagement_sheet,
        ["Статус", "Количество"],
        ((_enum_to_value(item.status), item.count) for item in stats.engagements),
    )

    initiator_sheet = wb.create_sheet(title="Сотрудничества по инициатору")
    _write_table(
        initiator_sheet,
        ["Инициатор", "Количество"],
        (
            (_enum_to_value(item.initiator), item.count)
            for item in stats.engagements_by_initiator
        ),
    )

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_education_stats_report(stats: EducationStatsResponse) -> bytes:
    wb = _build_workbook()
    _summary_sheet(
        wb,
        "Сводка",
        [
            ("Активные стажировки", stats.internships.active, None),
            ("Запланированные стажировки", stats.internships.planned, None),
            ("Завершенные стажировки", stats.internships.completed, None),
            ("Всего участников", stats.participants_total, None),
            ("Компании-партнеры", stats.partner_companies, None),
            ("Загрузка мест", stats.capacity_utilization, "0.00%"),
            ("Среднее время набора (дни)", stats.average_recruitment_days, "0.00"),
            ("Средний курс участников", stats.average_participant_course, "0.00"),
            ("Опубликованные стажировки", stats.published_internships, None),
        ],
    )

    participants_sheet = wb.create_sheet(title="Статусы участников")
    _write_table(
        participants_sheet,
        ["Статус", "Количество"],
        ((_enum_to_value(item.status), item.count) for item in stats.participants_by_status),
    )

    engagement_sheet = wb.create_sheet(title="Сотрудничества по статусу")
    _write_table(
        engagement_sheet,
        ["Статус", "Количество"],
        (
            (_enum_to_value(item.status), item.count)
            for item in stats.engagement_status_breakdown
        ),
    )

    initiator_sheet = wb.create_sheet(title="Сотрудничества по инициатору")
    _write_table(
        initiator_sheet,
        ["Инициатор", "Количество"],
        (
            (_enum_to_value(item.initiator), item.count)
            for item in stats.engagement_initiator_breakdown
        ),
    )

    invites_sheet = wb.create_sheet(title="Приглашения")
    _write_table(
        invites_sheet,
        ["Статус", "Количество"],
        (
            ("Активные", stats.invite_activity.active),
            ("Использованные", stats.invite_activity.used),
            ("Истекшие", stats.invite_activity.expired),
        ),
    )

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_admin_stats_report(stats: AdminStatsResponse) -> bytes:
    wb = _build_workbook()
    _summary_sheet(
        wb,
        "Сводка",
        [
            ("Активные соискатели", stats.active_applicants, None),
            ("Активные компании", stats.active_companies, None),
            ("Активные образовательные организации", stats.active_educations, None),
            ("Активные стажировки", stats.active_internships, None),
            ("Уровень трудоустройства", stats.employment_rate, "0.00%"),
            (
                "Средняя длительность стажировок (дни)",
                stats.average_internship_duration_days,
                "0.00",
            ),
            ("Рост компаний", stats.company_growth_percent, "0.00%"),
            (
                "Средняя заполняемость стажировок",
                stats.average_internship_fill_rate,
                "0.00%",
            ),
            (
                "Заявок на вакансию",
                stats.application_per_vacancy_ratio,
                "0.00",
            ),
        ],
    )

    internship_series_sheet = wb.create_sheet(title="Стажировки по месяцам")
    _write_table(
        internship_series_sheet,
        ["Месяц", "Количество"],
        ((point.month, point.count) for point in stats.internship_series),
        column_formats=("yyyy-mm", None),
    )

    vacancy_series_sheet = wb.create_sheet(title="Вакансии по месяцам")
    _write_table(
        vacancy_series_sheet,
        ["Месяц", "Количество"],
        ((point.month, point.count) for point in stats.vacancy_series),
        column_formats=("yyyy-mm", None),
    )

    engagement_sheet = wb.create_sheet(title="Статусы взаимодействий")
    _write_table(
        engagement_sheet,
        ["Статус", "Количество"],
        ((_enum_to_value(item.status), item.count) for item in stats.engagement_status_breakdown),
    )

    invites_sheet = wb.create_sheet(title="Приглашения")
    _write_table(
        invites_sheet,
        ["Статус", "Количество"],
        (
            ("Активные", stats.invite_summary.active),
            ("Использованные", stats.invite_summary.used),
            ("Истекшие", stats.invite_summary.expired),
        ),
    )

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_report_filename(prefix: str, extension: str = "xlsx") -> str:
    timestamp = datetime.utcnow().strftime("%Y-%m-%d")
    return f"{prefix}-{timestamp}.{extension}"
